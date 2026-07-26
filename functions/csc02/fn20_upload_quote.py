"""
fn20_upload_quote.py — CSC-02 견적서 엑셀 업로드 처리

교수님 요청 반영:
  1) 부품번호(P/N)로 매칭 — 품명은 고정, 명칭 정정이 필요할 때만 갱신
  2) 유로 단가는 최신값으로 갱신, 원화 기록은 price_history에 이력으로 보존
  3) 매칭 안 되는 신규 부품은 자동 INSERT하지 않고 후보 목록으로만 반환
     (담당자가 검토 후 별도로 등록하도록 — 지난 수작업 검증 프로세스와 동일한 안전장치)

⚠️ DB 선행 작업 필요 (2026-07-14 반영 완료):
   ALTER TABLE components ADD COLUMN IF NOT EXISTS nomenclature_kr text;
   CREATE TABLE IF NOT EXISTS price_history (...);
   → docs/schema_price_mapping_feature.sql 참고
"""

import io
from datetime import date, datetime, timezone

import openpyxl

import functions.db as _db
from functions.csc03.fn8_exchange_rate import get_exchange_rate

# 업로드 양식 헤더 (견적서_업로드양식.xlsx 와 1:1 대응)
EXPECTED_HEADERS = [
    "부품번호(P/N)*", "부품명(영문)", "부품명(한글)",
    "유로단가(EUR)*", "견적일자", "공급처", "비고",
]


def _find_header_row(ws) -> int:
    """양식 상단에 안내문(범례)이 있으므로, 실제 헤더 행을 찾아서 반환."""
    for row in ws.iter_rows(min_row=1, max_row=30):
        values = [c.value for c in row]
        if values and values[0] == EXPECTED_HEADERS[0]:
            return row[0].row
    raise ValueError(
        "업로드 양식의 헤더 행을 찾을 수 없습니다. "
        "견적서_업로드양식.xlsx 형식을 그대로 사용해주세요."
    )


def _parse_quote_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def upload_quote(file_bytes: bytes, source_file: str) -> dict:
    """
    견적서 엑셀(bytes)을 파싱해 components / price_history 에 반영.

    Args:
        file_bytes  : 업로드된 xlsx 파일의 바이트
        source_file : 원본 파일명 (이력 기록용)

    Returns:
        {
            "matched"        : int,   매칭되어 갱신된 행 수
            "unmatched"      : list,  매칭 안 된 부품번호 후보 목록 (신규 등록 검토용)
            "updated_parts"  : list,  실제 갱신된 부품 요약 (part_id, part_number, old/new price)
            "errors"         : list,  행별 처리 중 발생한 오류
        }

    Raises:
        ValueError: 헤더 형식이 다르거나, 필수값(P/N·유로단가) 누락 시
    """
    supabase = _db.get_client()

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header_row = _find_header_row(ws)

    # 최신 환율 조회 (원화환산액 계산용) — 없으면 None 으로 두고 원화 계산은 생략
    try:
        rate_info = get_exchange_rate(currency_code="EUR", base_currency="KRW")
        current_rate = rate_info["exchange_rate"]
    except ValueError:
        current_rate = None

    matched = 0
    unmatched = []
    updated_parts = []
    errors = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        part_number = row[0]
        if part_number is None or str(part_number).strip() == "":
            continue  # 빈 행 스킵

        part_number = str(part_number).strip()
        nomenclature_en = (str(row[1]).strip() if row[1] not in (None, "") else None)
        nomenclature_kr = (str(row[2]).strip() if row[2] not in (None, "") else None)
        unit_price_eur = row[3]
        quote_date = _parse_quote_date(row[4])
        supplier_name = (str(row[5]).strip() if row[5] not in (None, "") else None)

        if unit_price_eur is None:
            errors.append({"part_number": part_number, "error": "유로단가(EUR)가 비어 있습니다"})
            continue

        try:
            unit_price_eur = float(unit_price_eur)
        except (TypeError, ValueError):
            errors.append({"part_number": part_number, "error": f"유로단가 형식 오류: {unit_price_eur!r}"})
            continue

        # 1) 부품 매칭 (part_number 기준)
        comp = (
            supabase.table("components")
            .select("id, unit_price_eur, nomenclature, nomenclature_kr")
            .eq("part_number", part_number)
            .eq("is_deleted", False)
            .execute()
        ).data

        if not comp:
            unmatched.append({
                "part_number": part_number,
                "nomenclature_en": nomenclature_en,
                "nomenclature_kr": nomenclature_kr,
                "unit_price_eur": unit_price_eur,
                "quote_date": quote_date,
                "supplier": supplier_name,
            })
            continue

        part_id = comp[0]["id"]
        old_price = comp[0].get("unit_price_eur")

        # 2) 공급처 매칭 (이름으로 조회, 없으면 None 유지 — 신규 공급처 등록은 별도 처리)
        supplier_id = None
        if supplier_name:
            sup = (
                supabase.table("suppliers")
                .select("id")
                .eq("supplier_name", supplier_name)
                .execute()
            ).data
            supplier_id = sup[0]["id"] if sup else None

        # 3) components 업데이트 — 품명은 값이 있을 때만, 유로단가는 항상 최신화
        update_payload = {"unit_price_eur": unit_price_eur}
        if nomenclature_en:
            update_payload["nomenclature"] = nomenclature_en
        if nomenclature_kr:
            update_payload["nomenclature_kr"] = nomenclature_kr

        unit_price_krw = round(unit_price_eur * current_rate) if current_rate else None
        if unit_price_krw is not None:
            update_payload["unit_price_krw"] = unit_price_krw

        supabase.table("components").update(update_payload).eq("id", part_id).execute()

        # 4) price_history 에 이력 INSERT (원화 기록 보존)
        supabase.table("price_history").insert({
            "part_id": part_id,
            "unit_price_eur": unit_price_eur,
            "exchange_rate_applied": current_rate,
            "unit_price_krw": unit_price_krw,
            "source_file": source_file,
            "quote_date": quote_date,
            "supplier_id": supplier_id,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }).execute()

        matched += 1
        updated_parts.append({
            "part_id": part_id,
            "part_number": part_number,
            "old_unit_price_eur": old_price,
            "new_unit_price_eur": unit_price_eur,
            "unit_price_krw": unit_price_krw,
        })

    return {
        "matched": matched,
        "unmatched": unmatched,
        "updated_parts": updated_parts,
        "errors": errors or None,
    }
